from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl import load_workbook

from app.core.ids import generate_id
from app.core.reference_data import (
    BUSINESS_LINE_OPTIONS,
    METRIC_LIFECYCLE_ACTIVE,
    REPORT_TYPE_OPTIONS,
    clean_metric_text,
)
from app.repositories.metrics import MetricRepository


class MetricLibraryImportService:
    _header_aliases = {
        "metric_code": "metric_code",
        "指标编码": "metric_code",
        "code": "metric_code",
        "metric_name": "metric_name",
        "指标名称": "metric_name",
        "中文指标名": "metric_name",
        "metric_name_en": "metric_name_en",
        "英文名": "metric_name_en",
        "category": "category",
        "指标分类": "category",
        "subcategory": "subcategory",
        "二级分类": "subcategory",
        "definition": "definition",
        "定义": "definition",
        "value_type": "value_type",
        "值类型": "value_type",
        "default_unit": "default_unit",
        "默认单位": "default_unit",
        "aliases": "aliases",
        "别名": "aliases",
        "同义词": "aliases",
        "report_types": "report_types",
        "适用报告类型": "report_types",
        "business_lines": "business_lines",
        "适用业务条线": "business_lines",
        "parent_metric_code": "parent_metric_code",
        "父级指标编码": "parent_metric_code",
        "sort_order": "sort_order",
        "排序": "sort_order",
        "formula_expression": "formula_expression",
        "公式表达式": "formula_expression",
        "depends_on_codes": "depends_on_codes",
        "依赖指标编码": "depends_on_codes",
    }

    _required_fields = ("metric_code", "metric_name")

    def __init__(self, metric_repository: MetricRepository) -> None:
        self.metric_repository = metric_repository

    def build_template(self) -> bytes:
        workbook = Workbook()
        template_sheet = workbook.active
        template_sheet.title = "metrics_template"

        headers = [
            "指标编码",
            "指标名称",
            "英文名",
            "指标分类",
            "二级分类",
            "定义",
            "值类型",
            "默认单位",
            "别名",
            "适用报告类型",
            "适用业务条线",
            "父级指标编码",
            "排序",
            "公式表达式",
            "依赖指标编码",
        ]
        template_sheet.append(headers)
        template_sheet.append(
            [
                "core_premium_income",
                "保险业务收入",
                "Premium income",
                "盈利能力",
                "保险收入",
                "保险业务形成的主营收入",
                "AMOUNT",
                "元",
                "保费收入,保险收入",
                "annual_report,semiannual_report",
                "group,life,pnc",
                "",
                "10",
                "",
                "",
            ]
        )

        guide_sheet = workbook.create_sheet("instructions")
        guide_rows = [
            ("字段", "是否必填", "说明", "允许值/示例"),
            ("指标编码", "是", "唯一业务编码，用于导入更新匹配", "core_premium_income"),
            ("指标名称", "是", "中文标准指标名", "保险业务收入"),
            ("英文名", "否", "英文名称", "Premium income"),
            ("指标分类", "否", "一级分类", "盈利能力"),
            ("二级分类", "否", "二级分类", "保险收入"),
            ("定义", "否", "指标定义说明", "保险业务形成的主营收入"),
            ("值类型", "否", "未填默认 AMOUNT", "AMOUNT / RATIO / COUNT"),
            ("默认单位", "否", "指标默认单位", "元 / % / 次"),
            ("别名", "否", "多个值可用逗号或分号分隔", "保费收入,保险收入"),
            (
                "适用报告类型",
                "否",
                "可留空，留空表示全部报告类型",
                ", ".join(REPORT_TYPE_OPTIONS.keys()),
            ),
            (
                "适用业务条线",
                "否",
                "可留空，留空表示全部业务条线",
                ", ".join(BUSINESS_LINE_OPTIONS.keys()),
            ),
            ("父级指标编码", "否", "用于建立指标层级", "core_revenue"),
            ("排序", "否", "整数，未填默认 0", "10"),
            ("公式表达式", "否", "展示用途的公式文本", "A + B - C"),
            ("依赖指标编码", "否", "多个编码可用逗号或分号分隔", "metric_a,metric_b"),
        ]
        for row in guide_rows:
            guide_sheet.append(row)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def preview(self, file_name: str, file_bytes: bytes) -> dict:
        rows = self._load_rows(file_name, file_bytes)
        normalized_rows, errors = self._normalize_rows(rows)
        return {
            "total_rows": len(rows),
            "valid_rows": len(normalized_rows),
            "errors": errors,
            "preview_items": [
                {
                    "row_number": row["row_number"],
                    "metric_code": row["metric_code"],
                    "metric_name": row["metric_name"],
                    "aliases": row["aliases"],
                    "category": row["category"],
                    "default_unit": row["default_unit"],
                    "report_types": row["report_types"],
                    "business_lines": row["business_lines"],
                }
                for row in normalized_rows[:30]
            ],
        }

    def import_rows(self, file_name: str, file_bytes: bytes) -> dict:
        rows = self._load_rows(file_name, file_bytes)
        normalized_rows, errors = self._normalize_rows(rows)

        metrics_created = 0
        metrics_updated = 0
        aliases_created = 0
        aliases_updated = 0

        code_to_canonical_id: dict[str, str] = {}
        pending_dependencies: list[dict] = []

        for row in normalized_rows:
            existing = self.metric_repository.get_by_code(row["metric_code"])
            canonical_metric_id = existing.canonical_metric_id if existing is not None else self._canonical_metric_id(row["metric_code"])
            parent_metric = self.metric_repository.get_by_code(row["parent_metric_code"]) if row.get("parent_metric_code") else None
            payload = {
                "canonical_metric_id": canonical_metric_id,
                "metric_code": row["metric_code"],
                "metric_name": row["metric_name"],
                "metric_name_en": row["metric_name_en"],
                "category": row["category"],
                "subcategory": row["subcategory"],
                "definition": row["definition"],
                "formula_expression": row["formula_expression"],
                "value_type": row["value_type"],
                "default_unit": row["default_unit"],
                "parent_canonical_metric_id": parent_metric.canonical_metric_id if parent_metric is not None else None,
                "hierarchy_depth": 1 if parent_metric is not None else 0,
                "sort_order": row["sort_order"],
                "applicable_scope": {"values": row["business_lines"]},
                "applicable_statement_types": {"values": row["report_types"]},
                "is_active": True,
                "lifecycle_status": METRIC_LIFECYCLE_ACTIVE,
                "version": (existing.version + 1) if existing is not None else 1,
            }
            _, created = self.metric_repository.upsert_metric_definition(payload)
            code_to_canonical_id[row["metric_code"]] = canonical_metric_id
            if created:
                metrics_created += 1
            else:
                metrics_updated += 1

            alias_texts = [row["metric_name"], *row["aliases"]]
            for alias_text in self._unique(alias_texts):
                alias_payload = {
                    "alias_id": self._alias_id(canonical_metric_id, alias_text),
                    "canonical_metric_id": canonical_metric_id,
                    "alias_text": alias_text,
                    "alias_lang": "zh-CN",
                    "company_id": None,
                    "report_type": None,
                    "valid_from_year": None,
                    "valid_to_year": None,
                    "confidence": 0.98,
                    "source": "core_import",
                }
                _, alias_created = self.metric_repository.upsert_alias(alias_payload)
                if alias_created:
                    aliases_created += 1
                else:
                    aliases_updated += 1

            pending_dependencies.append(
                {
                    "canonical_metric_id": canonical_metric_id,
                    "depends_on_codes": row["depends_on_codes"],
                }
            )

        for dependency in pending_dependencies:
            self.metric_repository.clear_dependencies(dependency["canonical_metric_id"])
            for index, depends_on_code in enumerate(dependency["depends_on_codes"], start=1):
                depends_on_id = code_to_canonical_id.get(depends_on_code)
                if depends_on_id is None:
                    depends_on_metric = self.metric_repository.get_by_code(depends_on_code)
                    depends_on_id = depends_on_metric.canonical_metric_id if depends_on_metric is not None else None
                if depends_on_id is None:
                    continue
                self.metric_repository.add_dependency(
                    {
                        "dependency_id": generate_id("mdep"),
                        "canonical_metric_id": dependency["canonical_metric_id"],
                        "depends_on_metric_id": depends_on_id,
                        "relation_type": "formula_input",
                        "expression_hint": None,
                        "sort_order": index,
                    }
                )

        self.metric_repository.commit()
        return {
            "total_rows": len(rows),
            "metrics_created": metrics_created,
            "metrics_updated": metrics_updated,
            "aliases_created": aliases_created,
            "aliases_updated": aliases_updated,
            "errors": errors,
        }

    def _load_rows(self, file_name: str, file_bytes: bytes) -> list[dict]:
        suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
        if suffix == "csv":
            return self._load_csv(file_bytes)
        if suffix in {"xlsx", "xlsm"}:
            return self._load_xlsx(file_bytes)
        raise ValueError("仅支持 CSV 或 XLSX 指标模板")

    def _load_csv(self, file_bytes: bytes) -> list[dict]:
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    def _load_xlsx(self, file_bytes: bytes) -> list[dict]:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        results: list[dict] = []
        for values in rows[1:]:
            if values is None or not any(value is not None and str(value).strip() for value in values):
                continue
            results.append(
                {
                    headers[index]: ("" if index >= len(values) or values[index] is None else str(values[index]).strip())
                    for index in range(len(headers))
                }
            )
        return results

    def _normalize_rows(self, rows: list[dict]) -> tuple[list[dict], list[str]]:
        normalized_rows: list[dict] = []
        errors: list[str] = []

        for index, row in enumerate(rows, start=2):
            normalized = self._normalize_row(row, index)
            if isinstance(normalized, str):
                errors.append(normalized)
                continue
            normalized_rows.append(normalized)

        return normalized_rows, errors

    def _normalize_row(self, row: dict, row_number: int) -> dict | str:
        mapped: dict[str, str] = {}
        for key, value in row.items():
            if key is None:
                continue
            normalized_key = self._header_aliases.get(str(key).strip())
            if not normalized_key:
                continue
            mapped[normalized_key] = "" if value is None else str(value).strip()

        missing_fields = [field for field in self._required_fields if not mapped.get(field)]
        if missing_fields:
            return f"第 {row_number} 行缺少必填字段: {', '.join(missing_fields)}"

        metric_code = mapped["metric_code"].strip()
        metric_name = clean_metric_text(mapped["metric_name"])
        if not metric_name:
            return f"第 {row_number} 行指标名称无效"

        aliases = [alias for alias in self._split_multi_value(mapped.get("aliases")) if clean_metric_text(alias)]
        report_types = self._normalize_report_types(mapped.get("report_types"))
        business_lines = self._normalize_business_lines(mapped.get("business_lines"))
        if isinstance(report_types, str):
            return f"第 {row_number} 行{report_types}"
        if isinstance(business_lines, str):
            return f"第 {row_number} 行{business_lines}"

        sort_order = self._normalize_int(mapped.get("sort_order"), default=0)
        if sort_order is None:
            return f"第 {row_number} 行排序值无效"

        return {
            "row_number": row_number,
            "metric_code": metric_code,
            "metric_name": metric_name,
            "metric_name_en": mapped.get("metric_name_en") or None,
            "category": mapped.get("category") or None,
            "subcategory": mapped.get("subcategory") or None,
            "definition": mapped.get("definition") or None,
            "formula_expression": mapped.get("formula_expression") or None,
            "value_type": mapped.get("value_type") or "AMOUNT",
            "default_unit": mapped.get("default_unit") or None,
            "aliases": [clean_metric_text(alias) for alias in aliases if clean_metric_text(alias)],
            "report_types": report_types,
            "business_lines": business_lines,
            "parent_metric_code": mapped.get("parent_metric_code") or None,
            "sort_order": sort_order,
            "depends_on_codes": self._split_multi_value(mapped.get("depends_on_codes")),
        }

    def _normalize_report_types(self, raw_value: str | None) -> list[str] | str:
        if not raw_value:
            return list(REPORT_TYPE_OPTIONS.keys())
        results: list[str] = []
        reverse = {label: code for code, label in REPORT_TYPE_OPTIONS.items()}
        for item in self._split_multi_value(raw_value):
            candidate = item.strip()
            code = candidate if candidate in REPORT_TYPE_OPTIONS else reverse.get(candidate)
            if not code:
                return f"报告类型不合法: {candidate}"
            results.append(code)
        return self._unique(results)

    def _normalize_business_lines(self, raw_value: str | None) -> list[str] | str:
        if not raw_value:
            return list(BUSINESS_LINE_OPTIONS.keys())
        results: list[str] = []
        reverse = {label: code for code, label in BUSINESS_LINE_OPTIONS.items()}
        for item in self._split_multi_value(raw_value):
            candidate = item.strip()
            code = candidate if candidate in BUSINESS_LINE_OPTIONS else reverse.get(candidate)
            if not code:
                return f"业务条线不合法: {candidate}"
            results.append(code)
        return self._unique(results)

    @staticmethod
    def _normalize_int(value: str | None, default: int = 0) -> int | None:
        if value is None or value == "":
            return default
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _split_multi_value(value: str | None) -> list[str]:
        if not value:
            return []
        normalized = str(value).replace("；", ";").replace("，", ",").replace("\n", ";")
        parts = []
        for token in normalized.split(";"):
            parts.extend(part.strip() for part in token.split(","))
        return [part for part in parts if part]

    @staticmethod
    def _unique(values: Iterable[str]) -> list[str]:
        seen: set[str] = set()
        results: list[str] = []
        for value in values:
            if value in seen:
                continue
            seen.add(value)
            results.append(value)
        return results

    @staticmethod
    def _canonical_metric_id(metric_code: str) -> str:
        digest = hashlib.sha1(metric_code.strip().lower().encode("utf-8")).hexdigest()
        return f"metric_core_{digest[:16]}"

    @staticmethod
    def _alias_id(canonical_metric_id: str, alias_text: str) -> str:
        digest = hashlib.sha1(f"{canonical_metric_id}|{alias_text.strip().lower()}".encode("utf-8")).hexdigest()
        return f"alias_core_{digest[:20]}"
