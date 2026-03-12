from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.config import settings
from app.core.ids import generate_id
from app.models.document import Document
from app.repositories.documents import DocumentRepository
from app.repositories.table_qa import TableQARepository
from app.services.document_outline import DocumentOutlineService


class TableQAService:
    _generic_titles = {
        "word-layout numeric table",
        "ocr-detected numeric table",
        "auto-detected numeric table",
    }

    def __init__(self, document_repository: DocumentRepository, table_qa_repository: TableQARepository) -> None:
        self.document_repository = document_repository
        self.table_qa_repository = table_qa_repository

    def list_document_qa(self, document_id: str) -> list[dict]:
        document = self.document_repository.get(document_id)
        if document is None:
            raise ValueError(f"document not found: {document_id}")

        tables = self.document_repository.list_tables(document_id)
        reviews = {review.table_id: review for review in self.table_qa_repository.list_by_document(document_id)}
        outline = DocumentOutlineService().build_outline(
            document=document,
            tables=[
                {
                    "table_id": table.table_id,
                    "table_title_raw": table.table_title_raw,
                    "table_title_norm": table.table_title_norm,
                    "page_start": table.page_start,
                    "page_end": table.page_end,
                }
                for table in tables
            ],
        )
        section_paths = self._build_section_path_map(outline)

        rows = []
        for table in sorted(tables, key=lambda item: (item.page_start, item.page_end, item.table_id)):
            review = reviews.get(table.table_id)
            rows.append(self._build_row(document, table, section_paths.get(table.table_id, []), review))
        return rows

    def save_review(self, document_id: str, table_id: str, manual_status: str | None, manual_note: str | None, reviewer: str | None) -> dict:
        document = self.document_repository.get(document_id)
        if document is None:
            raise ValueError(f"document not found: {document_id}")
        table = self.document_repository.get_table(document_id, table_id)
        if table is None:
            raise ValueError(f"table not found: {table_id}")
        review = self.table_qa_repository.upsert(
            {
                "qa_review_id": generate_id("tqa"),
                "document_id": document_id,
                "table_id": table_id,
                "manual_status": manual_status,
                "manual_note": manual_note,
                "reviewer": reviewer,
                "metadata_json": None,
            }
        )
        return {
            "qa_review_id": review.qa_review_id,
            "table_id": review.table_id,
            "manual_status": review.manual_status,
            "manual_note": review.manual_note,
            "reviewer": review.reviewer,
            "reviewed_at": review.reviewed_at,
        }

    def export_document_qa(self, document_id: str) -> Path:
        rows = self.list_document_qa(document_id)
        export_root = Path(settings.export_root) / "table_qa"
        export_root.mkdir(parents=True, exist_ok=True)
        export_path = export_root / f"{document_id}_table_qa.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "table_qa"

        headers = [
            "section_path",
            "page_start",
            "page_end",
            "table_id",
            "table_title",
            "parse_engine",
            "header_levels",
            "col_header_count",
            "row_count",
            "cell_count",
            "generic_title",
            "generic_headers",
            "merged_header_detected",
            "suspected_wrap_issue",
            "suspected_value_issue",
            "bbox_quality",
            "overall_status",
            "auto_flags",
            "viewer_url",
            "manual_status",
            "manual_note",
            "reviewer",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            values = [
                row["section_path"],
                row["page_start"],
                row["page_end"],
                row["table_id"],
                row["table_title"],
                row["parse_engine"],
                row["header_levels"],
                row["col_header_count"],
                row["row_count"],
                row["cell_count"],
                row["generic_title"],
                row["generic_headers"],
                row["merged_header_detected"],
                row["suspected_wrap_issue"],
                row["suspected_value_issue"],
                row["bbox_quality"],
                row["overall_status"],
                " | ".join(row["auto_flags"]),
                row["viewer_url"],
                row["manual_status"],
                row["manual_note"],
                row["reviewer"],
            ]
            sheet.append(values)
            link_cell = sheet.cell(row=sheet.max_row, column=headers.index("viewer_url") + 1)
            link_cell.hyperlink = row["viewer_url"]
            link_cell.style = "Hyperlink"

        for column in sheet.columns:
            letter = column[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 40)

        workbook.save(export_path)
        return export_path

    def _build_row(self, document: Document, table: Any, section_path: list[str], review: Any | None) -> dict:
        table_json = getattr(table, "table_json", {}) or {}
        parse_trace = getattr(table, "parse_trace_json", {}) or {}
        cells = table_json.get("cells", []) if isinstance(table_json, dict) else []
        row_headers = table_json.get("row_headers", []) if isinstance(table_json, dict) else []
        col_headers = table_json.get("col_headers", []) if isinstance(table_json, dict) else []
        title = getattr(table, "table_title_raw", None) or getattr(table, "table_title_norm", None) or "未命名表格"
        nonzero_bbox_cells = [cell for cell in cells if self._has_nonzero_bbox(cell.get("bbox", []))]
        header_levels = max((len(cell.get("col_path", [])) for cell in cells if isinstance(cell.get("col_path"), list)), default=1)
        generic_title = self._is_generic_title(title)
        generic_headers = sum(1 for header in col_headers if str(header).startswith("value_") or str(header) == "current_period")
        merged_header_detected = header_levels > 1 or len(set(col_headers)) < len(col_headers)
        suspected_wrap_issue = self._suspected_wrap_issue(row_headers)
        suspected_value_issue = generic_headers > 0 or any(self._is_suspect_value(str(cell.get("value_raw", ""))) for cell in cells)
        bbox_quality = self._bbox_quality(nonzero_bbox_cells, cells, parse_trace)
        auto_flags = []
        if generic_title:
            auto_flags.append("GENERIC_TITLE")
        if generic_headers:
            auto_flags.append("GENERIC_HEADERS")
        if merged_header_detected:
            auto_flags.append("MERGED_HEADERS")
        if suspected_wrap_issue:
            auto_flags.append("WRAP_ISSUE")
        if suspected_value_issue:
            auto_flags.append("VALUE_ISSUE")
        if bbox_quality != "strong":
            auto_flags.append("WEAK_BBOX")

        if generic_title and bbox_quality == "weak":
            overall_status = "FAIL"
        elif auto_flags:
            overall_status = "REVIEW"
        else:
            overall_status = "PASS"

        return {
            "document_id": document.document_id,
            "document_label": document.document_label,
            "table_id": table.table_id,
            "section_path": " / ".join(section_path) if section_path else "未归类",
            "page_start": table.page_start,
            "page_end": table.page_end,
            "table_title": title,
            "parse_engine": getattr(table, "parse_engine", None) or "unknown",
            "header_levels": header_levels,
            "col_header_count": len(col_headers),
            "row_count": len(row_headers),
            "cell_count": len(cells),
            "generic_title": generic_title,
            "generic_headers": generic_headers,
            "merged_header_detected": merged_header_detected,
            "suspected_wrap_issue": suspected_wrap_issue,
            "suspected_value_issue": suspected_value_issue,
            "bbox_quality": bbox_quality,
            "overall_status": overall_status,
            "auto_flags": auto_flags,
            "viewer_url": f"/documents/{document.document_id}/viewer?page={table.page_start}&table_id={table.table_id}",
            "manual_status": getattr(review, "manual_status", None),
            "manual_note": getattr(review, "manual_note", None),
            "reviewer": getattr(review, "reviewer", None),
            "reviewed_at": getattr(review, "reviewed_at", None),
        }

    def _build_section_path_map(self, outline: list[dict]) -> dict[str, list[str]]:
        mapping: dict[str, list[str]] = {}

        def walk(nodes: list[dict], path: list[str]) -> None:
            for node in nodes:
                title = str(node.get("title") or "").strip()
                next_path = [*path, title] if title else path
                for table in node.get("tables", []):
                    table_id = table.get("table_id")
                    if table_id:
                        mapping[str(table_id)] = next_path
                walk(node.get("children", []), next_path)

        walk(outline, [])
        return mapping

    @classmethod
    def _is_generic_title(cls, title: str) -> bool:
        normalized = title.strip().lower()
        if normalized in cls._generic_titles:
            return True
        return normalized.startswith("第 ") and normalized.endswith(" 页表格")

    @staticmethod
    def _has_nonzero_bbox(bbox: list[Any]) -> bool:
        return isinstance(bbox, list) and len(bbox) == 4 and any(float(value) != 0.0 for value in bbox)

    @staticmethod
    def _bbox_quality(nonzero_bbox_cells: list[dict], cells: list[dict], parse_trace: dict) -> str:
        if not cells:
            return "weak"
        ratio = len(nonzero_bbox_cells) / max(len(cells), 1)
        table_bbox = parse_trace.get("bbox") if isinstance(parse_trace, dict) else None
        if ratio >= 0.85 and isinstance(table_bbox, list) and len(table_bbox) == 4:
            return "strong"
        if ratio >= 0.4 or (isinstance(table_bbox, list) and len(table_bbox) == 4):
            return "medium"
        return "weak"

    @staticmethod
    def _suspected_wrap_issue(row_headers: list[Any]) -> bool:
        if not row_headers:
            return False
        long_headers = [header for header in row_headers if len(str(header).replace(" ", "")) >= 24]
        trailing_symbols = [header for header in row_headers if str(header).rstrip().endswith(("及", "和", "、", "-", "（"))]
        return (len(long_headers) / max(len(row_headers), 1)) > 0.25 or bool(trailing_symbols)

    @staticmethod
    def _is_suspect_value(value_raw: str) -> bool:
        compact = value_raw.strip().replace(" ", "")
        return compact in {"", "-", "--", "...", "…"}
